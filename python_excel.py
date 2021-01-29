import openpyxl
import os

print(os.getcwd())

wb = openpyxl.load_workbook('example.xlsx')

print(f'Our workbook type -> [{str(type(wb))}]')

#Getting the sheets' names
print(f"Sheets name -> {wb.sheetnames}")

#Getting the second sheet
sheet_1 = wb['Sheet1']
print(sheet_1.title)

#We can access to the cell value once we have the sheet
#also we can access to the row, column and value
a_cell = sheet_1['A1']
print(a_cell.value)
print(f'Row [{a_cell.row}] Column [{a_cell.column}] Value [{a_cell.value}]')

#it's a little bit tricky access to the cell value by a letter
#so we can use the cell() method and pass it the row and column number
#the row/column first value is 1, not 0.

for i in range(1, 20, 2):
    if(sheet_1.cell(i, 2).value is None):
        print(f'Breaking the loop at {i} index because the cell value is null.')
        break
    print(f'{i} - Cell value [{sheet_1.cell(row =i, column=2).value}]')

#we can determine the size of the sheet by the max_row and max_column attributes
print(f'Sheet max column: {sheet_1.max_column}')
print(f'Sheet max row: {sheet_1.max_row}')

#you can slice worksheet objects to get all the cell objects in a row, column
#or rectangular area of the spreadsheet

print('Slicing the worksheet')
print('---------------------')
for row in sheet_1['A1':'C3']:
    for cell in row:
        print(cell.coordinate, cell.value)
    print('------ END OF ROW -------')